import openpyxl as oxl
import pandas as pd
import numpy as np
from thefuzz import fuzz
from thefuzz import process as fzproc


def get_end_row_index(sheet, n_footers):
    end_row = len(list(sheet.values)) - n_footers
    return end_row


def remove_blank_rows(sheet, start_row, n_footers):
    
    end_row = get_end_row_index(sheet, n_footers)

    # Iterate over rows in reverse order (from bottom to top)
    for row in range(end_row, start_row - 1, -1):
        is_blank_row = True
        # Check if all cells in the row are blank
        for col in range(1, sheet.max_column + 1):  # Assuming checking all columns
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip() != '':
                is_blank_row = False
                break
        
        # If the row is blank, delete it
        if is_blank_row:
            sheet.delete_rows(row, 1)
    
    return sheet


def get_column_from_sheet(sheet, column, start_row, n_footers, substr: None | str = None):
    end_row = get_end_row_index(sheet, n_footers)

    col_values = []

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column)
        cell_value = cell.value
        col_values.append(cell_value)
    
    if substr is not None and type(substr) == str:
        col_values = [value for value in col_values if substr in str(value)]
    
    return col_values


def change_indent_based_on_series(sheet, series, start_row, n_footers, delta_indent):
    end_row = get_end_row_index(sheet, n_footers)

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=1)
        cell_value = cell.value
        # indent_level = cell.alignment.indent

        if cell_value in series.values:
            current_indent = cell.alignment.indent or 0
            new_indent = current_indent + delta_indent
            cell.alignment = cell.alignment.copy(indent=new_indent)
        
    return sheet


def insert_indent_status_column(sheet, start_row, n_footers):
    
    end_row = get_end_row_index(sheet, n_footers)

    # Insert a new column to the left of column A
    sheet.insert_cols(1)
    
    # Iterate through rows and set values based on indent level
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=1)  # Cell in column B (2nd column)
        indent_level = sheet.cell(row=row, column=2).alignment.indent

        if indent_level == 0:
            cell.value = "R"
        elif indent_level == 1:
            cell.value = "P"
        elif indent_level == 2:
            cell.value = "M"
        else:
            cell.value = ""
    
    return sheet

def get_huc_names(df: pd.DataFrame, choices: list, name_col = "Name", col_to_check = "City Class", col_value = "HUC", scorer = fuzz.WRatio, score_cutoff = 90, pateros: bool = True) -> pd.Series:
    if pateros:
        choices.append("PATEROS")
    
    has_value = df[col_to_check] == col_value
    if pateros:
        is_pateros = df[name_col] == "Pateros"
        df_filtered = df.copy(deep=True)[has_value | is_pateros]
    else:
        df_filtered = df.copy(deep=True)[has_value]

    df_filtered["match"] = df_filtered[name_col].apply(
        lambda x: fzproc.extractOne(
            x,
            choices,
            scorer=scorer,
            score_cutoff=score_cutoff
        )
    ).str[0]
    
    return df_filtered["match"]